# app.py - Modern Business Management System v8.0
import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog
import database as db
import datetime
import shutil
import os
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from receipt_generator import ReceiptGenerator

# Set appearance mode and color theme
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

# ========== HELPER FUNCTIONS ==========

def format_date(date_str):
    """Convert date from YYYY-MM-DD to dd/mm/yyyy format"""
    try:
        if not date_str:
            return ""
        # Parse the date string
        date_obj = datetime.datetime.strptime(str(date_str), '%Y-%m-%d')
        # Format as dd/mm/yyyy (4-digit year)
        return date_obj.strftime('%d/%m/%Y')
    except:
        # If parsing fails, return original
        return str(date_str)

# ========== DIALOG WINDOWS ==========

class EditTransactionWindow(ctk.CTkToplevel):
    """Pop-up window for editing transactions"""

    def __init__(self, master, transaction_id):
        super().__init__(master)
        self.master_app = master
        self.transaction_id = transaction_id

        self.title("Επεξεργασία Συναλλαγής")
        self.geometry("500x400")
        self.transient(master)
        self.grab_set()

        # Get transaction details
        _id, current_notes, current_status = db.get_transaction_details(self.transaction_id)

        # Main container
        main_frame = ctk.CTkFrame(self)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)

        # Title
        title_label = ctk.CTkLabel(
            main_frame,
            text=f"Επεξεργασία Συναλλαγής #{transaction_id}",
            font=ctk.CTkFont(size=18, weight="bold")
        )
        title_label.pack(pady=(0, 20))

        # Status
        status_label = ctk.CTkLabel(main_frame, text="Κατάσταση Πληρωμής:", font=ctk.CTkFont(weight="bold"))
        status_label.pack(pady=(10, 5), anchor="w")

        self.status_var = ctk.StringVar(value=current_status)
        self.status_menu = ctk.CTkOptionMenu(
            main_frame,
            variable=self.status_var,
            values=["Εκκρεμεί", "Πληρώθηκε"]
        )
        self.status_menu.pack(fill="x", pady=(0, 10))

        # Notes
        notes_label = ctk.CTkLabel(main_frame, text="Σχόλια / Παρατηρήσεις:", font=ctk.CTkFont(weight="bold"))
        notes_label.pack(pady=(10, 5), anchor="w")

        self.notes_textbox = ctk.CTkTextbox(main_frame, height=150)
        self.notes_textbox.pack(fill="both", expand=True, pady=(0, 15))
        self.notes_textbox.insert("1.0", current_notes if current_notes else "")

        # Buttons
        button_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        button_frame.pack(fill="x", pady=(10, 0))

        save_btn = ctk.CTkButton(
            button_frame,
            text="💾 Αποθήκευση",
            command=self.save_changes,
            height=40,
            font=ctk.CTkFont(size=14)
        )
        save_btn.pack(side="left", fill="x", expand=True, padx=(0, 5))

        cancel_btn = ctk.CTkButton(
            button_frame,
            text="✖ Ακύρωση",
            command=self.destroy,
            height=40,
            fg_color="gray",
            font=ctk.CTkFont(size=14)
        )
        cancel_btn.pack(side="right", fill="x", expand=True, padx=(5, 0))

    def save_changes(self):
        new_status = self.status_var.get()
        new_notes = self.notes_textbox.get("1.0", "end-1c").strip()

        db.update_transaction(self.transaction_id, new_status, new_notes)
        messagebox.showinfo("Επιτυχία", "Οι αλλαγές αποθηκεύτηκαν επιτυχώς.", parent=self)

        # Refresh main app views
        if hasattr(self.master_app, 'refresh_main_table'):
            self.master_app.refresh_main_table()
        if hasattr(self.master_app, 'refresh_customer_view'):
            self.master_app.refresh_customer_view()

        self.destroy()


class CustomerProfileWindow(ctk.CTkToplevel):
    """Customer profile view and edit window"""

    def __init__(self, master, customer_name):
        super().__init__(master)
        self.master_app = master
        self.customer_name = customer_name

        self.title(f"Προφίλ Πελάτη - {customer_name}")
        self.geometry("900x700")

        # Get customer ID and details
        self.customer_id = db.get_customer_id_by_name(customer_name)
        if not self.customer_id:
            messagebox.showerror("Σφάλμα", f"Δεν βρέθηκε ο πελάτης: {customer_name}")
            self.destroy()
            return

        customer_details = db.get_customer_details(self.customer_id)

        # Create scrollable main frame
        self.main_frame = ctk.CTkScrollableFrame(self)
        self.main_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # Header
        header_frame = ctk.CTkFrame(self.main_frame)
        header_frame.pack(fill="x", pady=(0, 20))

        title_label = ctk.CTkLabel(
            header_frame,
            text=f"👤 {customer_name}",
            font=ctk.CTkFont(size=24, weight="bold")
        )
        title_label.pack(side="left", padx=20, pady=15)

        # Customer Details Section
        details_frame = ctk.CTkFrame(self.main_frame)
        details_frame.pack(fill="both", padx=10, pady=(0, 20))

        details_title = ctk.CTkLabel(
            details_frame,
            text="📋 Στοιχεία Πελάτη",
            font=ctk.CTkFont(size=16, weight="bold")
        )
        details_title.pack(pady=15, padx=15, anchor="w")

        # Fields container
        fields_frame = ctk.CTkFrame(details_frame, fg_color="transparent")
        fields_frame.pack(fill="both", padx=15, pady=(0, 15))

        # Left column
        left_col = ctk.CTkFrame(fields_frame, fg_color="transparent")
        left_col.pack(side="left", fill="both", expand=True, padx=(0, 10))

        # Name
        self.create_field(left_col, "Ονοματεπώνυμο:", customer_details[1] if customer_details else "", "name_entry")

        # Email
        self.create_field(left_col, "Email:", customer_details[2] if customer_details else "", "email_entry")

        # Phone
        self.create_field(left_col, "Τηλέφωνο:", customer_details[3] if customer_details else "", "phone_entry")

        # Tax ID
        self.create_field(left_col, "ΑΦΜ:", customer_details[4] if customer_details else "", "tax_id_entry")

        # Address
        self.create_field(left_col, "Διεύθυνση:", customer_details[5] if customer_details else "", "address_entry")

        # Right column
        right_col = ctk.CTkFrame(fields_frame, fg_color="transparent")
        right_col.pack(side="right", fill="both", expand=True, padx=(10, 0))

        # Work Info
        self.create_field(right_col, "Εργασία:", customer_details[6] if customer_details else "", "work_entry")

        # TAXIS Credentials Section
        taxis_label = ctk.CTkLabel(right_col, text="🔐 Κωδικοί TAXIS Net", font=ctk.CTkFont(size=14, weight="bold"))
        taxis_label.pack(pady=(15, 10), anchor="w")

        # TAXIS Username
        self.create_field(right_col, "Username:", customer_details[7] if customer_details else "", "taxis_user_entry")

        # TAXIS Password
        self.create_field(right_col, "Password:", customer_details[8] if customer_details else "", "taxis_pass_entry", show="*")

        # Notes (full width)
        notes_label = ctk.CTkLabel(details_frame, text="📝 Σημειώσεις:", font=ctk.CTkFont(weight="bold"))
        notes_label.pack(pady=(10, 5), padx=15, anchor="w")

        self.notes_textbox = ctk.CTkTextbox(details_frame, height=80)
        self.notes_textbox.pack(fill="x", padx=15, pady=(0, 15))
        self.notes_textbox.insert("1.0", customer_details[9] if customer_details and customer_details[9] else "")

        # Save button
        save_btn = ctk.CTkButton(
            details_frame,
            text="💾 Αποθήκευση Στοιχείων",
            command=self.save_customer_details,
            height=40,
            font=ctk.CTkFont(size=14)
        )
        save_btn.pack(fill="x", padx=15, pady=(0, 15))

        # Transactions Section
        trans_frame = ctk.CTkFrame(self.main_frame)
        trans_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        trans_title = ctk.CTkLabel(
            trans_frame,
            text="💰 Ιστορικό Συναλλαγών",
            font=ctk.CTkFont(size=16, weight="bold")
        )
        trans_title.pack(pady=15, padx=15, anchor="w")

        # Summary
        records = db.get_transactions_by_customer(customer_name)
        total_paid = sum(r[4] for r in records if r[5] == 'Πληρώθηκε')
        total_unpaid = sum(r[4] for r in records if r[5] != 'Πληρώθηκε')

        summary_frame = ctk.CTkFrame(trans_frame, fg_color="transparent")
        summary_frame.pack(fill="x", padx=15, pady=(0, 10))

        paid_label = ctk.CTkLabel(
            summary_frame,
            text=f"✅ Πληρωμένα: {total_paid:.2f} €",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#4ade80"
        )
        paid_label.pack(side="left", padx=(0, 20))

        unpaid_label = ctk.CTkLabel(
            summary_frame,
            text=f"❌ Οφειλές: {total_unpaid:.2f} €",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#f87171"
        )
        unpaid_label.pack(side="left")

        # Transactions tree
        tree_frame = ctk.CTkFrame(trans_frame)
        tree_frame.pack(fill="both", expand=True, padx=15, pady=(0, 15))

        columns = ("ID", "Υπηρεσία", "Ημερομηνία", "Ποσό", "Κατάσταση")
        self.trans_tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=8)

        for col in columns:
            self.trans_tree.heading(col, text=col)

        self.trans_tree.column("ID", width=50, anchor="center")
        self.trans_tree.column("Υπηρεσία", width=200)
        self.trans_tree.column("Ημερομηνία", width=100, anchor="center")
        self.trans_tree.column("Ποσό", width=100, anchor="e")
        self.trans_tree.column("Κατάσταση", width=100, anchor="center")

        self.trans_tree.tag_configure('paid', background='#166534', foreground='white')
        self.trans_tree.tag_configure('unpaid', background='#991b1b', foreground='white')

        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.trans_tree.yview)
        self.trans_tree.configure(yscrollcommand=scrollbar.set)

        self.trans_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Load transactions
        for record in records:
            trans_id, service, notes, date, cost, status = record
            tag = 'paid' if status == 'Πληρώθηκε' else 'unpaid'
            formatted_date = format_date(date)
            self.trans_tree.insert("", "end", values=(trans_id, service, formatted_date, f"{cost:.2f} €", status), tags=(tag,))

        # Transaction actions
        actions_frame = ctk.CTkFrame(trans_frame, fg_color="transparent")
        actions_frame.pack(fill="x", padx=15, pady=(0, 15))

        edit_trans_btn = ctk.CTkButton(
            actions_frame,
            text="✏️ Επεξεργασία",
            command=self.edit_selected_transaction,
            height=35
        )
        edit_trans_btn.pack(side="left", padx=(0, 5))

        delete_trans_btn = ctk.CTkButton(
            actions_frame,
            text="🗑️ Διαγραφή",
            command=self.delete_selected_transaction,
            fg_color="#dc2626",
            hover_color="#991b1b",
            height=35
        )
        delete_trans_btn.pack(side="left", padx=(5, 5))

        receipt_btn = ctk.CTkButton(
            actions_frame,
            text="🧾 Απόδειξη",
            command=self.generate_receipt,
            height=35
        )
        receipt_btn.pack(side="left", padx=(5, 0))

    def create_field(self, parent, label_text, value, attr_name, show=None):
        """Helper to create labeled entry fields"""
        label = ctk.CTkLabel(parent, text=label_text, font=ctk.CTkFont(weight="bold"))
        label.pack(pady=(10, 2), anchor="w")

        entry = ctk.CTkEntry(parent, height=35)
        if show:
            entry.configure(show=show)
        entry.pack(fill="x", pady=(0, 5))
        entry.insert(0, value if value else "")

        setattr(self, attr_name, entry)

    def save_customer_details(self):
        """Save updated customer details"""
        name = self.name_entry.get().strip()
        email = self.email_entry.get().strip()
        phone = self.phone_entry.get().strip()
        tax_id = self.tax_id_entry.get().strip()
        address = self.address_entry.get().strip()
        work_info = self.work_entry.get().strip()
        taxis_user = self.taxis_user_entry.get().strip()
        taxis_pass = self.taxis_pass_entry.get().strip()
        notes = self.notes_textbox.get("1.0", "end-1c").strip()

        if not name:
            messagebox.showerror("Σφάλμα", "Το όνομα του πελάτη είναι υποχρεωτικό.", parent=self)
            return

        try:
            db.update_customer_details(
                self.customer_id, name, email, phone, tax_id,
                address, work_info, taxis_user, taxis_pass, notes
            )
            messagebox.showinfo("Επιτυχία", "Τα στοιχεία του πελάτη ενημερώθηκαν επιτυχώς.", parent=self)

            # Update title if name changed
            if name != self.customer_name:
                self.customer_name = name
                self.title(f"Προφίλ Πελάτη - {name}")

        except Exception as e:
            messagebox.showerror("Σφάλμα", f"Αποτυχία ενημέρωσης: {str(e)}", parent=self)

    def edit_selected_transaction(self):
        """Edit selected transaction"""
        selected = self.trans_tree.selection()
        if not selected:
            messagebox.showwarning("Προσοχή", "Παρακαλώ επιλέξτε μια συναλλαγή.", parent=self)
            return

        trans_id = self.trans_tree.item(selected[0])['values'][0]
        EditTransactionWindow(self, trans_id)

    def delete_selected_transaction(self):
        """Delete selected transaction"""
        selected = self.trans_tree.selection()
        if not selected:
            messagebox.showwarning("Προσοχή", "Παρακαλώ επιλέξτε μια συναλλαγή.", parent=self)
            return

        trans_id = self.trans_tree.item(selected[0])['values'][0]

        if messagebox.askyesno("Επιβεβαίωση Διαγραφής",
                               f"Είστε σίγουροι ότι θέλετε να διαγράψετε τη συναλλαγή #{trans_id};\n\nΗ ενέργεια δεν μπορεί να αναιρεθεί.",
                               parent=self):
            db.delete_transaction(trans_id)
            messagebox.showinfo("Επιτυχία", "Η συναλλαγή διαγράφηκε επιτυχώς.", parent=self)

            # Refresh view
            self.trans_tree.delete(selected[0])

            # Refresh main app
            if hasattr(self.master_app, 'refresh_main_table'):
                self.master_app.refresh_main_table()

    def generate_receipt(self):
        """Generate receipt for selected transaction"""
        selected = self.trans_tree.selection()
        if not selected:
            messagebox.showwarning("Προσοχή", "Παρακαλώ επιλέξτε μια συναλλαγή.", parent=self)
            return

        trans_id = self.trans_tree.item(selected[0])['values'][0]
        service = self.trans_tree.item(selected[0])['values'][1]
        date_formatted = self.trans_tree.item(selected[0])['values'][2]  # This is already in dd/mm/yy format
        amount_str = self.trans_tree.item(selected[0])['values'][3]
        amount = float(amount_str.replace(' €', '').replace(',', '.'))

        # Show receipt options dialog
        ReceiptOptionsWindow(self, trans_id, self.customer_name, service, amount, date_formatted)


class ReceiptOptionsWindow(ctk.CTkToplevel):
    """Receipt generation options window"""

    def __init__(self, master, trans_id, customer_name, service, amount, date):
        super().__init__(master)
        self.trans_id = trans_id
        self.customer_name = customer_name
        self.service = service
        self.amount = amount
        self.date = date

        self.title("Δημιουργία Απόδειξης")
        self.geometry("600x700")
        self.transient(master)
        self.grab_set()

        # Main frame
        main_frame = ctk.CTkScrollableFrame(self)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)

        # Title
        title_label = ctk.CTkLabel(
            main_frame,
            text="🧾 Δημιουργία Απόδειξης",
            font=ctk.CTkFont(size=20, weight="bold")
        )
        title_label.pack(pady=(0, 20))

        # Receipt Type
        type_frame = ctk.CTkFrame(main_frame)
        type_frame.pack(fill="x", pady=(0, 15))

        type_label = ctk.CTkLabel(type_frame, text="Τύπος Απόδειξης:", font=ctk.CTkFont(weight="bold"))
        type_label.pack(pady=(15, 5), padx=15, anchor="w")

        self.receipt_type = ctk.StringVar(value="payment")

        payment_radio = ctk.CTkRadioButton(
            type_frame,
            text="Απόδειξη Πληρωμής",
            variable=self.receipt_type,
            value="payment"
        )
        payment_radio.pack(padx=20, pady=5, anchor="w")

        collection_radio = ctk.CTkRadioButton(
            type_frame,
            text="Απόδειξη Είσπραξης",
            variable=self.receipt_type,
            value="collection"
        )
        collection_radio.pack(padx=20, pady=(0, 15), anchor="w")

        # Company Settings
        settings_frame = ctk.CTkFrame(main_frame)
        settings_frame.pack(fill="x", pady=(0, 15))

        settings_label = ctk.CTkLabel(settings_frame, text="📄 Στοιχεία Εταιρείας:", font=ctk.CTkFont(weight="bold", size=14))
        settings_label.pack(pady=(15, 10), padx=15, anchor="w")

        # Load existing company settings
        existing_settings = db.get_company_settings()

        # Company Name
        self.create_settings_field(settings_frame, "Όνομα Εταιρείας:",
                                   existing_settings[0] if existing_settings else "", "company_name_entry")

        # Company Address
        self.create_settings_field(settings_frame, "Διεύθυνση:",
                                   existing_settings[3] if existing_settings else "", "company_address_entry")

        # Company Phone
        self.create_settings_field(settings_frame, "Τηλέφωνο:",
                                   existing_settings[4] if existing_settings else "", "company_phone_entry")

        # Company Email
        self.create_settings_field(settings_frame, "Email:",
                                   existing_settings[5] if existing_settings else "", "company_email_entry")

        # Company Tax ID
        self.create_settings_field(settings_frame, "ΑΦΜ:",
                                   existing_settings[6] if existing_settings else "", "company_tax_entry")

        # Logo
        logo_label = ctk.CTkLabel(settings_frame, text="Logo Εταιρείας:", font=ctk.CTkFont(weight="bold"))
        logo_label.pack(pady=(10, 5), padx=15, anchor="w")

        logo_frame = ctk.CTkFrame(settings_frame, fg_color="transparent")
        logo_frame.pack(fill="x", padx=15, pady=(0, 10))

        self.logo_path = ctk.StringVar(value=existing_settings[1] if existing_settings and existing_settings[1] else "")
        self.logo_label = ctk.CTkLabel(
            logo_frame,
            text=os.path.basename(self.logo_path.get()) if self.logo_path.get() else "Κανένα αρχείο",
            text_color="gray"
        )
        self.logo_label.pack(side="left", padx=(0, 10))

        logo_btn = ctk.CTkButton(logo_frame, text="Επιλογή Logo", command=self.select_logo, width=120)
        logo_btn.pack(side="left")

        # Signature
        sig_label = ctk.CTkLabel(settings_frame, text="Υπογραφή:", font=ctk.CTkFont(weight="bold"))
        sig_label.pack(pady=(10, 5), padx=15, anchor="w")

        sig_frame = ctk.CTkFrame(settings_frame, fg_color="transparent")
        sig_frame.pack(fill="x", padx=15, pady=(0, 15))

        self.signature_path = ctk.StringVar(value=existing_settings[2] if existing_settings and existing_settings[2] else "")
        self.sig_label = ctk.CTkLabel(
            sig_frame,
            text=os.path.basename(self.signature_path.get()) if self.signature_path.get() else "Κανένα αρχείο",
            text_color="gray"
        )
        self.sig_label.pack(side="left", padx=(0, 10))

        sig_btn = ctk.CTkButton(sig_frame, text="Επιλογή Υπογραφής", command=self.select_signature, width=120)
        sig_btn.pack(side="left")

        # Save settings checkbox
        self.save_settings_var = ctk.IntVar(value=1)
        save_settings_check = ctk.CTkCheckBox(
            settings_frame,
            text="Αποθήκευση στοιχείων για μελλοντική χρήση",
            variable=self.save_settings_var
        )
        save_settings_check.pack(padx=15, pady=(0, 15), anchor="w")

        # Comments Section
        comments_frame = ctk.CTkFrame(main_frame)
        comments_frame.pack(fill="x", pady=(0, 15))

        comments_label = ctk.CTkLabel(comments_frame, text="💬 Σχόλια για την Απόδειξη:", font=ctk.CTkFont(weight="bold", size=14))
        comments_label.pack(pady=(15, 10), padx=15, anchor="w")

        self.receipt_comments_textbox = ctk.CTkTextbox(comments_frame, height=80)
        self.receipt_comments_textbox.pack(fill="x", padx=15, pady=(0, 15))

        # Generate Button
        generate_btn = ctk.CTkButton(
            main_frame,
            text="📄 Δημιουργία Απόδειξης",
            command=self.generate_receipt,
            height=45,
            font=ctk.CTkFont(size=15, weight="bold")
        )
        generate_btn.pack(fill="x", pady=(15, 0))

    def create_settings_field(self, parent, label_text, value, attr_name):
        """Helper to create settings entry fields"""
        label = ctk.CTkLabel(parent, text=label_text, font=ctk.CTkFont(weight="bold"))
        label.pack(pady=(5, 2), padx=15, anchor="w")

        entry = ctk.CTkEntry(parent, height=32)
        entry.pack(fill="x", padx=15, pady=(0, 5))
        entry.insert(0, value if value else "")

        setattr(self, attr_name, entry)

    def select_logo(self):
        """Select logo file"""
        filepath = filedialog.askopenfilename(
            title="Επιλογή Logo",
            filetypes=[("Image files", "*.png *.jpg *.jpeg *.gif"), ("All files", "*.*")]
        )
        if filepath:
            self.logo_path.set(filepath)
            self.logo_label.configure(text=os.path.basename(filepath), text_color="white")

    def select_signature(self):
        """Select signature file"""
        filepath = filedialog.askopenfilename(
            title="Επιλογή Υπογραφής",
            filetypes=[("Image files", "*.png *.jpg *.jpeg *.gif"), ("All files", "*.*")]
        )
        if filepath:
            self.signature_path.set(filepath)
            self.sig_label.configure(text=os.path.basename(filepath), text_color="white")

    def generate_receipt(self):
        """Generate the receipt PDF"""
        company_name = self.company_name_entry.get().strip()
        company_address = self.company_address_entry.get().strip()
        company_phone = self.company_phone_entry.get().strip()
        company_email = self.company_email_entry.get().strip()
        company_tax = self.company_tax_entry.get().strip()

        if not company_name:
            messagebox.showwarning("Προσοχή", "Παρακαλώ εισάγετε το όνομα της εταιρείας.", parent=self)
            return

        # Save settings if checkbox is checked
        if self.save_settings_var.get():
            db.update_company_settings(
                company_name,
                self.logo_path.get(),
                self.signature_path.get(),
                company_address,
                company_phone,
                company_email,
                company_tax
            )

        # Ask where to save
        default_filename = f"Apoδειξη_{self.trans_id}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        output_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=default_filename,
            title="Αποθήκευση Απόδειξης"
        )

        if not output_path:
            return

        # Get comments from textbox
        receipt_comments = self.receipt_comments_textbox.get("1.0", "end-1c").strip()

        # Create receipt generator
        generator = ReceiptGenerator(
            company_name=company_name,
            company_address=company_address,
            company_phone=company_phone,
            company_email=company_email,
            company_tax_id=company_tax,
            logo_path=self.logo_path.get() if self.logo_path.get() else None,
            signature_path=self.signature_path.get() if self.signature_path.get() else None
        )

        try:
            if self.receipt_type.get() == "payment":
                generator.generate_payment_receipt(
                    output_path,
                    f"#{self.trans_id}",
                    self.customer_name,
                    self.amount,
                    self.service,
                    payment_date=self.date,
                    notes=receipt_comments
                )
            else:
                generator.generate_collection_receipt(
                    output_path,
                    f"#{self.trans_id}",
                    self.customer_name,
                    self.amount,
                    self.service,
                    collection_date=self.date,
                    notes=receipt_comments
                )

            messagebox.showinfo("Επιτυχία", f"Η απόδειξη δημιουργήθηκε επιτυχώς!\n\n{output_path}", parent=self)

            # Ask if user wants to open the file
            if messagebox.askyesno("Άνοιγμα Αρχείου", "Θέλετε να ανοίξετε την απόδειξη;", parent=self):
                os.startfile(output_path)

            self.destroy()

        except Exception as e:
            messagebox.showerror("Σφάλμα", f"Αποτυχία δημιουργίας απόδειξης:\n{str(e)}", parent=self)


# ========== MAIN APPLICATION ==========

class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Σύστημα Διαχείρισης Έργων v8.0 - Modern Edition")
        self.geometry("1400x800")

        # Initialize database
        db.connect_db()

        # State variables
        self.current_customer_records = []
        self.current_customer_name = None

        # Create tab view
        self.tab_view = ctk.CTkTabview(self)
        self.tab_view.pack(expand=True, fill="both", padx=15, pady=15)

        # Create tabs
        self.main_tab = self.tab_view.add("🏠 Αρχική")
        self.customers_tab = self.tab_view.add("👥 Πελάτες")
        self.services_tab = self.tab_view.add("⚙️ Υπηρεσίες")
        self.import_tab = self.tab_view.add("📤 Εισαγωγή")
        self.log_tab = self.tab_view.add("📋 Ιστορικό")

        # Build tabs
        self.create_main_tab()
        self.create_customers_tab()
        self.create_services_tab()
        self.create_import_tab()
        self.create_log_tab()

        # Set default tab
        self.tab_view.set("🏠 Αρχική")

    # ========== MAIN TAB (Home) ==========

    def create_main_tab(self):
        """Create the main home tab with transaction entry and list"""
        self.main_tab.grid_columnconfigure(1, weight=1)
        self.main_tab.grid_rowconfigure(0, weight=1)

        # Left Panel - New Transaction Form
        left_panel = ctk.CTkScrollableFrame(self.main_tab, width=380)
        left_panel.grid(row=0, column=0, padx=(10, 5), pady=10, sticky="ns")
        left_panel.grid_columnconfigure(0, weight=1)

        # Form Title
        form_title = ctk.CTkLabel(
            left_panel,
            text="➕ Νέα Καταχώρηση",
            font=ctk.CTkFont(size=22, weight="bold")
        )
        form_title.pack(pady=(10, 25))

        # Customer Name
        customer_label = ctk.CTkLabel(left_panel, text="Όνομα Πελάτη *", font=ctk.CTkFont(weight="bold"))
        customer_label.pack(pady=(0, 5), padx=20, anchor="w")

        self.customer_name_entry = ctk.CTkEntry(left_panel, height=40, placeholder_text="Εισάγετε όνομα πελάτη...")
        self.customer_name_entry.pack(fill="x", padx=20, pady=(0, 15))

        # Service
        service_label = ctk.CTkLabel(left_panel, text="Υπηρεσία *", font=ctk.CTkFont(weight="bold"))
        service_label.pack(pady=(0, 5), padx=20, anchor="w")

        self.service_var = ctk.StringVar(value="Επιλέξτε Υπηρεσία...")
        self.service_menu = ctk.CTkOptionMenu(left_panel, variable=self.service_var, values=[], height=40)
        self.service_menu.pack(fill="x", padx=20, pady=(0, 15))
        self.update_services_dropdown()

        # Notes
        notes_label = ctk.CTkLabel(left_panel, text="Σχόλια / Παρατηρήσεις", font=ctk.CTkFont(weight="bold"))
        notes_label.pack(pady=(0, 5), padx=20, anchor="w")

        self.notes_entry = ctk.CTkEntry(left_panel, height=40, placeholder_text="Προαιρετικό...")
        self.notes_entry.pack(fill="x", padx=20, pady=(0, 15))

        # Final Cost
        cost_label = ctk.CTkLabel(left_panel, text="Τελικό Κόστος (με ΦΠΑ) *", font=ctk.CTkFont(weight="bold"))
        cost_label.pack(pady=(0, 5), padx=20, anchor="w")

        self.final_cost_var = ctk.StringVar()
        self.final_cost_var.trace_add("write", self.calculate_vat)
        self.cost_final_entry = ctk.CTkEntry(left_panel, textvariable=self.final_cost_var, height=40, placeholder_text="0.00")
        self.cost_final_entry.pack(fill="x", padx=20, pady=(0, 15))

        # Pre-VAT Cost (readonly)
        prevat_label = ctk.CTkLabel(left_panel, text="Κόστος προ ΦΠΑ (αυτόματο)", font=ctk.CTkFont(weight="bold"))
        prevat_label.pack(pady=(0, 5), padx=20, anchor="w")

        self.cost_pre_vat_entry = ctk.CTkEntry(left_panel, height=40, state="readonly")
        self.cost_pre_vat_entry.pack(fill="x", padx=20, pady=(0, 15))

        # Payment Status
        status_label = ctk.CTkLabel(left_panel, text="Κατάσταση Πληρωμής *", font=ctk.CTkFont(weight="bold"))
        status_label.pack(pady=(0, 5), padx=20, anchor="w")

        self.status_var = ctk.StringVar(value="Εκκρεμεί")
        self.status_menu = ctk.CTkOptionMenu(
            left_panel,
            variable=self.status_var,
            values=["Εκκρεμεί", "Πληρώθηκε"],
            height=40
        )
        self.status_menu.pack(fill="x", padx=20, pady=(0, 15))

        # File Attachment
        attachment_label = ctk.CTkLabel(left_panel, text="Επισυναπτόμενο Αρχείο", font=ctk.CTkFont(weight="bold"))
        attachment_label.pack(pady=(0, 5), padx=20, anchor="w")

        self.attachment_path = ctk.StringVar()
        self.attachment_label = ctk.CTkLabel(
            left_panel,
            text="Κανένα αρχείο επιλεγμένο",
            text_color="gray"
        )
        self.attachment_label.pack(pady=(0, 5), padx=20)

        attach_btn = ctk.CTkButton(
            left_panel,
            text="📎 Επισύναψη Αρχείου",
            command=self.select_file,
            height=35
        )
        attach_btn.pack(fill="x", padx=20, pady=(0, 25))

        # Submit Button
        submit_btn = ctk.CTkButton(
            left_panel,
            text="✅ Προσθήκη Εγγραφής",
            command=self.add_transaction,
            height=50,
            font=ctk.CTkFont(size=16, weight="bold")
        )
        submit_btn.pack(fill="x", padx=20, pady=(0, 20))

        # Right Panel - Transactions List
        right_panel = ctk.CTkFrame(self.main_tab)
        right_panel.grid(row=0, column=1, padx=(5, 10), pady=10, sticky="nsew")
        right_panel.grid_columnconfigure(0, weight=1)
        right_panel.grid_rowconfigure(2, weight=1)

        # Transactions Title
        trans_title = ctk.CTkLabel(
            right_panel,
            text="📊 Συναλλαγές",
            font=ctk.CTkFont(size=22, weight="bold")
        )
        trans_title.grid(row=0, column=0, pady=(15, 10), padx=20, sticky="w")

        # Filter Frame
        filter_frame = ctk.CTkFrame(right_panel, fg_color="transparent")
        filter_frame.grid(row=1, column=0, padx=20, pady=(0, 10), sticky="ew")

        filter_label = ctk.CTkLabel(filter_frame, text="Φίλτρο:", font=ctk.CTkFont(weight="bold"))
        filter_label.pack(side="left", padx=(0, 10))

        self.filter_var = ctk.StringVar(value="Όλα")
        filter_menu = ctk.CTkOptionMenu(
            filter_frame,
            variable=self.filter_var,
            values=["Όλα", "Εκκρεμεί", "Πληρώθηκε"],
            command=self.refresh_main_table,
            width=150
        )
        filter_menu.pack(side="left")

        # Transactions Treeview
        tree_frame = ctk.CTkFrame(right_panel)
        tree_frame.grid(row=2, column=0, padx=20, pady=(0, 10), sticky="nsew")

        columns = ("ID", "Πελάτης", "Υπηρεσία", "Σχόλια", "Ημερομηνία", "Ποσό", "Κατάσταση")
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings")

        for col in columns:
            self.tree.heading(col, text=col)

        self.tree.column("ID", width=50, anchor="center")
        self.tree.column("Πελάτης", width=150)
        self.tree.column("Υπηρεσία", width=200)
        self.tree.column("Σχόλια", width=200)
        self.tree.column("Ημερομηνία", width=100, anchor="center")
        self.tree.column("Ποσό", width=100, anchor="e")
        self.tree.column("Κατάσταση", width=100, anchor="center")

        self.tree.tag_configure('paid', background='#166534', foreground='white')
        self.tree.tag_configure('unpaid', background='#991b1b', foreground='white')

        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)

        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        self.tree.bind("<Double-1>", self.on_tree_double_click)

        # Action Buttons
        action_frame = ctk.CTkFrame(right_panel, fg_color="transparent")
        action_frame.grid(row=3, column=0, padx=20, pady=(0, 15), sticky="ew")

        edit_btn = ctk.CTkButton(
            action_frame,
            text="✏️ Επεξεργασία",
            command=self.edit_selected_transaction,
            height=35,
            width=140
        )
        edit_btn.pack(side="left", padx=(0, 5))

        delete_btn = ctk.CTkButton(
            action_frame,
            text="🗑️ Διαγραφή",
            command=self.delete_selected_transaction,
            fg_color="#dc2626",
            hover_color="#991b1b",
            height=35,
            width=140
        )
        delete_btn.pack(side="left", padx=(5, 0))

        # Refresh table
        self.refresh_main_table()

    def calculate_vat(self, *args):
        """Calculate pre-VAT cost automatically"""
        try:
            final_cost = float(self.final_cost_var.get())
            pre_vat_cost = final_cost / 1.24

            # Update readonly field
            self.cost_pre_vat_entry.configure(state="normal")
            self.cost_pre_vat_entry.delete(0, 'end')
            self.cost_pre_vat_entry.insert(0, f"{pre_vat_cost:.2f}")
            self.cost_pre_vat_entry.configure(state="readonly")
        except (ValueError, Exception):
            self.cost_pre_vat_entry.configure(state="normal")
            self.cost_pre_vat_entry.delete(0, 'end')
            self.cost_pre_vat_entry.configure(state="readonly")

    def select_file(self):
        """Select file attachment"""
        filepath = filedialog.askopenfilename(title="Επιλογή Αρχείου")
        if filepath:
            self.attachment_path.set(filepath)
            self.attachment_label.configure(text=os.path.basename(filepath), text_color="white")

    def add_transaction(self):
        """Add new transaction"""
        customer_name = self.customer_name_entry.get().strip()
        service_name = self.service_var.get()
        notes = self.notes_entry.get().strip()
        cost_final = self.cost_final_entry.get()
        cost_pre_vat = self.cost_pre_vat_entry.get()
        status = self.status_var.get()

        # Validation
        if not all([customer_name, cost_final, service_name not in ["Επιλέξτε Υπηρεσία...", "Προσθέστε υπηρεσίες", "-"]]):
            messagebox.showerror("Σφάλμα", "Παρακαλώ συμπληρώστε όλα τα υποχρεωτικά πεδία (*).")
            return

        try:
            cost_final_float = float(cost_final)
            cost_pre_vat_float = float(cost_pre_vat)
        except ValueError:
            messagebox.showerror("Σφάλμα", "Το κόστος πρέπει να είναι αριθμός.")
            return

        # Get or create customer
        customer_id = db.get_customer_by_name(customer_name)
        if not customer_id:
            db.add_customer(customer_name)
            customer_id = db.get_customer_by_name(customer_name)

        # Get service ID
        service_id = {name: sid for sid, name in db.get_services()}.get(service_name)

        # Handle attachment
        final_attachment_path = ""
        original_path = self.attachment_path.get()
        if original_path:
            filename = f"{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}_{os.path.basename(original_path)}"
            final_attachment_path = os.path.join(db.ATTACHMENTS_DIR, filename)
            shutil.copy(original_path, final_attachment_path)

        # Add transaction
        db.add_transaction(
            customer_id, service_id, notes,
            datetime.date.today().strftime('%Y-%m-%d'),
            cost_pre_vat_float, cost_final_float, status, final_attachment_path
        )

        # Log the action
        db.add_audit_log(
            "INSERT", "transactions", 0,
            f"Νέα συναλλαγή: {customer_name} - {service_name} - {cost_final_float:.2f}€",
            "", ""
        )

        messagebox.showinfo("Επιτυχία", "Η εγγραφή προστέθηκε επιτυχώς!")
        self.clear_form()
        self.refresh_main_table()

    def clear_form(self):
        """Clear the transaction form"""
        self.customer_name_entry.delete(0, 'end')
        self.notes_entry.delete(0, 'end')
        self.cost_final_entry.delete(0, 'end')
        self.attachment_path.set("")
        self.attachment_label.configure(text="Κανένα αρχείο επιλεγμένο", text_color="gray")

    def refresh_main_table(self, filter_choice=None):
        """Refresh the main transactions table"""
        if filter_choice is None:
            filter_choice = self.filter_var.get()

        # Clear existing items
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Get records
        records = db.get_all_transactions(filter_choice)

        for record in records:
            trans_id, customer, service, notes, date, amount, status = record
            tag = 'paid' if status == 'Πληρώθηκε' else 'unpaid'
            formatted_date = format_date(date)
            self.tree.insert("", "end", values=(trans_id, customer, service, notes, formatted_date, f"{amount:.2f} €", status), tags=(tag,))

    def on_tree_double_click(self, event):
        """Handle double-click on transaction"""
        selected = self.tree.selection()
        if selected:
            trans_id = self.tree.item(selected[0])['values'][0]
            EditTransactionWindow(self, trans_id)

    def edit_selected_transaction(self):
        """Edit selected transaction"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Προσοχή", "Παρακαλώ επιλέξτε μια συναλλαγή.")
            return

        trans_id = self.tree.item(selected[0])['values'][0]
        EditTransactionWindow(self, trans_id)

    def delete_selected_transaction(self):
        """Delete selected transaction"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Προσοχή", "Παρακαλώ επιλέξτε μια συναλλαγή.")
            return

        trans_id = self.tree.item(selected[0])['values'][0]

        if messagebox.askyesno("Επιβεβαίωση Διαγραφής",
                               f"Είστε σίγουροι ότι θέλετε να διαγράψετε τη συναλλαγή #{trans_id}?\n\nΗ ενέργεια δεν μπορεί να αναιρεθεί."):
            db.delete_transaction(trans_id)
            messagebox.showinfo("Επιτυχία", "Η συναλλαγή διαγράφηκε επιτυχώς.")
            self.refresh_main_table()

    # ========== CUSTOMERS TAB ==========

    def create_customers_tab(self):
        """Create the customers management tab"""
        self.customers_tab.grid_columnconfigure(0, weight=1)
        self.customers_tab.grid_rowconfigure(1, weight=1)

        # Search Frame
        search_frame = ctk.CTkFrame(self.customers_tab)
        search_frame.grid(row=0, column=0, padx=20, pady=20, sticky="ew")

        search_title = ctk.CTkLabel(
            search_frame,
            text="🔍 Αναζήτηση Πελάτη",
            font=ctk.CTkFont(size=20, weight="bold")
        )
        search_title.pack(pady=(15, 10), padx=20, anchor="w")

        search_subtitle = ctk.CTkLabel(
            search_frame,
            text="Αναζητήστε με οποιοδήποτε μέρος του ονόματος (π.χ. Νίκος, Κούκος, Νι, Κου...)",
            text_color="gray"
        )
        search_subtitle.pack(pady=(0, 10), padx=20, anchor="w")

        search_input_frame = ctk.CTkFrame(search_frame, fg_color="transparent")
        search_input_frame.pack(fill="x", padx=20, pady=(0, 15))

        self.customer_search_entry = ctk.CTkEntry(
            search_input_frame,
            height=45,
            placeholder_text="Πληκτρολογήστε όνομα πελάτη..."
        )
        self.customer_search_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        self.customer_search_entry.bind("<KeyRelease>", self.update_customer_suggestions)

        search_btn = ctk.CTkButton(
            search_input_frame,
            text="🔍 Αναζήτηση",
            command=self.search_customer,
            height=45,
            width=150,
            font=ctk.CTkFont(size=14, weight="bold")
        )
        search_btn.pack(side="right")

        # Suggestions listbox
        self.customer_suggestions_frame = ctk.CTkFrame(search_frame)
        self.customer_suggestions = []

        # Results Frame (will show customer profile when selected)
        self.customer_results_frame = ctk.CTkFrame(self.customers_tab)
        self.customer_results_frame.grid(row=1, column=0, padx=20, pady=(0, 20), sticky="nsew")

        # Initial empty state
        empty_label = ctk.CTkLabel(
            self.customer_results_frame,
            text="👆 Αναζητήστε έναν πελάτη για να δείτε τα στοιχεία του",
            font=ctk.CTkFont(size=16),
            text_color="gray"
        )
        empty_label.pack(expand=True)

    def update_customer_suggestions(self, event):
        """Update customer search suggestions (fuzzy search)"""
        search_term = self.customer_search_entry.get().strip()

        # Clear existing suggestions
        for widget in self.customer_suggestions_frame.winfo_children():
            widget.destroy()
        self.customer_suggestions_frame.pack_forget()

        if len(search_term) < 2:
            return

        # Get fuzzy search results
        results = db.fuzzy_search_customers(search_term)

        if results:
            self.customer_suggestions_frame.pack(fill="x", padx=20, pady=(0, 15))

            for customer_id, customer_name in results[:5]:  # Show top 5
                suggestion_btn = ctk.CTkButton(
                    self.customer_suggestions_frame,
                    text=customer_name,
                    command=lambda name=customer_name: self.select_customer_suggestion(name),
                    fg_color="transparent",
                    hover_color=("gray70", "gray30"),
                    anchor="w",
                    height=35
                )
                suggestion_btn.pack(fill="x", padx=5, pady=2)

    def select_customer_suggestion(self, customer_name):
        """Select a customer from suggestions"""
        self.customer_search_entry.delete(0, 'end')
        self.customer_search_entry.insert(0, customer_name)

        # Hide suggestions
        for widget in self.customer_suggestions_frame.winfo_children():
            widget.destroy()
        self.customer_suggestions_frame.pack_forget()

        # Open customer profile
        self.search_customer()

    def search_customer(self):
        """Search and display customer profile"""
        customer_name = self.customer_search_entry.get().strip()

        if not customer_name:
            messagebox.showwarning("Προσοχή", "Παρακαλώ εισάγετε όνομα πελάτη.")
            return

        # Check if customer exists
        customer_id = db.get_customer_id_by_name(customer_name)
        if not customer_id:
            # Try fuzzy search
            results = db.fuzzy_search_customers(customer_name)
            if results and len(results) == 1:
                customer_name = results[0][1]
                self.customer_search_entry.delete(0, 'end')
                self.customer_search_entry.insert(0, customer_name)
            else:
                messagebox.showinfo("Δεν βρέθηκε", f"Ο πελάτης '{customer_name}' δεν βρέθηκε στη βάση δεδομένων.")
                return

        # Open customer profile window
        CustomerProfileWindow(self, customer_name)

    def refresh_customer_view(self):
        """Refresh customer view (called after edits)"""
        if self.current_customer_name:
            self.search_customer()

    # ========== SERVICES TAB ==========

    def create_services_tab(self):
        """Create the services management tab"""
        self.services_tab.grid_columnconfigure(0, weight=1)
        self.services_tab.grid_rowconfigure(1, weight=1)

        # Add Service Frame
        add_frame = ctk.CTkFrame(self.services_tab)
        add_frame.grid(row=0, column=0, padx=20, pady=20, sticky="ew")

        add_title = ctk.CTkLabel(
            add_frame,
            text="➕ Προσθήκη Νέας Υπηρεσίας",
            font=ctk.CTkFont(size=20, weight="bold")
        )
        add_title.pack(pady=(15, 15), padx=20, anchor="w")

        input_frame = ctk.CTkFrame(add_frame, fg_color="transparent")
        input_frame.pack(fill="x", padx=20, pady=(0, 15))

        self.new_service_entry = ctk.CTkEntry(
            input_frame,
            height=45,
            placeholder_text="Όνομα υπηρεσίας..."
        )
        self.new_service_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))

        add_btn = ctk.CTkButton(
            input_frame,
            text="✅ Προσθήκη",
            command=self.add_new_service,
            height=45,
            width=150,
            font=ctk.CTkFont(size=14, weight="bold")
        )
        add_btn.pack(side="right")

        # Services List Frame
        list_frame = ctk.CTkFrame(self.services_tab)
        list_frame.grid(row=1, column=0, padx=20, pady=(0, 20), sticky="nsew")
        list_frame.grid_columnconfigure(0, weight=1)
        list_frame.grid_rowconfigure(1, weight=1)

        list_title = ctk.CTkLabel(
            list_frame,
            text="📋 Λίστα Υπηρεσιών",
            font=ctk.CTkFont(size=20, weight="bold")
        )
        list_title.grid(row=0, column=0, pady=(15, 10), padx=20, sticky="w")

        # Treeview for services
        tree_container = ctk.CTkFrame(list_frame)
        tree_container.grid(row=1, column=0, padx=20, pady=(0, 10), sticky="nsew")

        columns = ("ID", "Όνομα Υπηρεσίας")
        self.service_list_tree = ttk.Treeview(tree_container, columns=columns, show="headings")

        self.service_list_tree.heading("ID", text="ID")
        self.service_list_tree.heading("Όνομα Υπηρεσίας", text="Όνομα Υπηρεσίας")

        self.service_list_tree.column("ID", width=80, anchor="center")
        self.service_list_tree.column("Όνομα Υπηρεσίας", width=400)

        scrollbar = ttk.Scrollbar(tree_container, orient="vertical", command=self.service_list_tree.yview)
        self.service_list_tree.configure(yscrollcommand=scrollbar.set)

        self.service_list_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Delete Button
        delete_btn = ctk.CTkButton(
            list_frame,
            text="🗑️ Διαγραφή Επιλεγμένης Υπηρεσίας",
            command=self.delete_selected_service,
            fg_color="#dc2626",
            hover_color="#991b1b",
            height=40,
            font=ctk.CTkFont(size=14)
        )
        delete_btn.grid(row=2, column=0, padx=20, pady=(0, 15), sticky="ew")

        self.refresh_service_list()

    def add_new_service(self):
        """Add a new service"""
        service_name = self.new_service_entry.get().strip()

        if not service_name:
            messagebox.showwarning("Προσοχή", "Το όνομα της υπηρεσίας δεν μπορεί να είναι κενό.")
            return

        db.add_service(service_name)
        db.add_audit_log("INSERT", "services", 0, f"Νέα υπηρεσία: {service_name}", "", "")

        self.new_service_entry.delete(0, 'end')
        self.refresh_service_list()
        self.update_services_dropdown()

        messagebox.showinfo("Επιτυχία", f"Η υπηρεσία '{service_name}' προστέθηκε επιτυχώς!")

    def delete_selected_service(self):
        """Delete selected service"""
        selected = self.service_list_tree.selection()
        if not selected:
            messagebox.showwarning("Προσοχή", "Παρακαλώ επιλέξτε μια υπηρεσία για διαγραφή.")
            return

        service_id = self.service_list_tree.item(selected[0])['values'][0]
        service_name = self.service_list_tree.item(selected[0])['values'][1]

        if messagebox.askyesno("Επιβεβαίωση Διαγραφής",
                               f"Είστε σίγουροι ότι θέλετε να διαγράψετε την υπηρεσία '{service_name}'?\n\nΟι υπάρχουσες συναλλαγές θα δείχνουν 'Διαγραμμένη Υπηρεσία'."):
            db.delete_service(service_id)
            db.add_audit_log("DELETE", "services", service_id, f"Διαγραφή υπηρεσίας: {service_name}", "", "")

            self.refresh_service_list()
            self.update_services_dropdown()

            messagebox.showinfo("Επιτυχία", "Η υπηρεσία διαγράφηκε επιτυχώς.")

    def refresh_service_list(self):
        """Refresh the services list"""
        for item in self.service_list_tree.get_children():
            self.service_list_tree.delete(item)

        for service in db.get_services():
            self.service_list_tree.insert("", "end", values=service)

    def update_services_dropdown(self):
        """Update the services dropdown in main tab"""
        services = db.get_services()
        service_names = [s[1] for s in services] or ["-"]

        self.service_menu.configure(values=service_names)

        if service_names[0] != "-":
            self.service_var.set(service_names[0])
        else:
            self.service_var.set("Προσθέστε υπηρεσίες")

    # ========== IMPORT TAB ==========

    def create_import_tab(self):
        """Create the batch import tab"""
        self.import_tab.grid_columnconfigure(0, weight=1)
        self.import_tab.grid_rowconfigure(2, weight=1)

        # Info Frame
        info_frame = ctk.CTkFrame(self.import_tab)
        info_frame.grid(row=0, column=0, padx=20, pady=20, sticky="ew")

        title_label = ctk.CTkLabel(
            info_frame,
            text="📤 Μαζική Εισαγωγή Δεδομένων",
            font=ctk.CTkFont(size=22, weight="bold")
        )
        title_label.pack(pady=(15, 10), padx=20, anchor="w")

        step1_label = ctk.CTkLabel(
            info_frame,
            text="Βήμα 1: Κατεβάστε το πρότυπο αρχείο Excel",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        step1_label.pack(pady=(10, 5), padx=20, anchor="w")

        download_btn = ctk.CTkButton(
            info_frame,
            text="⬇️ Λήψη Προτύπου Excel",
            command=self.download_template,
            height=40,
            font=ctk.CTkFont(size=14)
        )
        download_btn.pack(fill="x", padx=20, pady=(0, 15))

        step2_label = ctk.CTkLabel(
            info_frame,
            text="Βήμα 2: Συμπληρώστε το αρχείο και εισάγετέ το",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        step2_label.pack(pady=(10, 5), padx=20, anchor="w")

        import_btn = ctk.CTkButton(
            info_frame,
            text="📥 Εισαγωγή από Excel",
            command=self.import_from_excel,
            height=40,
            font=ctk.CTkFont(size=14),
            fg_color="#059669",
            hover_color="#047857"
        )
        import_btn.pack(fill="x", padx=20, pady=(0, 15))

        # Log Frame
        log_title_frame = ctk.CTkFrame(self.import_tab)
        log_title_frame.grid(row=1, column=0, padx=20, pady=(0, 10), sticky="ew")

        log_title = ctk.CTkLabel(
            log_title_frame,
            text="📋 Αποτελέσματα Εισαγωγής",
            font=ctk.CTkFont(size=18, weight="bold")
        )
        log_title.pack(pady=15, padx=20, anchor="w")

        # Log Textbox
        self.import_log_textbox = ctk.CTkTextbox(self.import_tab, wrap="word")
        self.import_log_textbox.grid(row=2, column=0, padx=20, pady=(0, 20), sticky="nsew")
        self.import_log_textbox.insert("end", "Εδώ θα εμφανιστούν τα αποτελέσματα της διαδικασίας εισαγωγής...")
        self.import_log_textbox.configure(state="disabled")

    def download_template(self):
        """Download Excel template for batch import"""
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="Προτυπο_Εισαγωγης.xlsx",
            title="Αποθήκευση Προτύπου"
        )
        if not filepath:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"

            headers = [
                'Ονοματεπώνυμο Πελάτη', 'Υπηρεσία', 'Ημερομηνία (YYYY-MM-DD)',
                'Τελικό Κόστος (με ΦΠΑ)', 'Κατάσταση', 'Σχόλια'
            ]
            ws.append(headers)

            # Autofit columns
            for col_idx, header in enumerate(headers, 1):
                column_letter = get_column_letter(col_idx)
                ws.column_dimensions[column_letter].width = len(header) + 5

            wb.save(filepath)
            messagebox.showinfo("Επιτυχία", f"Το πρότυπο αποθηκεύτηκε επιτυχώς:\n{filepath}")

        except Exception as e:
            messagebox.showerror("Σφάλμα", f"Απέτυχε η δημιουργία του προτύπου:\n{e}")

    def import_from_excel(self):
        """Import transactions from Excel file"""
        filepath = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Επιλογή Αρχείου Excel"
        )
        if not filepath:
            return

        if not messagebox.askyesno("Επιβεβαίωση",
                                   "Είστε σίγουροι ότι θέλετε να ξεκινήσετε την εισαγωγή δεδομένων;"):
            return

        self.import_log_textbox.configure(state="normal")
        self.import_log_textbox.delete("1.0", "end")

        log = []
        success_count = 0
        fail_count = 0

        try:
            wb = load_workbook(filepath, data_only=True)
            ws = wb.active

            # Create service lookup dictionary
            available_services = {name.lower(): sid for sid, name in db.get_services()}

            # Process rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                try:
                    customer_name = str(row[0].value).strip() if row[0].value else None
                    service_name = str(row[1].value).strip() if row[1].value else None
                    date_val = row[2].value
                    final_cost = row[3].value
                    status = str(row[4].value).strip() if row[4].value else None
                    notes = str(row[5].value).strip() if row[5].value else ""

                    # Validation
                    if not all([customer_name, service_name, date_val, final_cost, status]):
                        raise ValueError("Λείπουν υποχρεωτικά δεδομένα")

                    # Validate service
                    service_id = available_services.get(service_name.lower())
                    if not service_id:
                        raise ValueError(f"Η υπηρεσία '{service_name}' δεν υπάρχει")

                    # Validate date
                    if isinstance(date_val, datetime.datetime):
                        transaction_date = date_val.strftime('%Y-%m-%d')
                    else:
                        transaction_date = str(date_val)
                        datetime.datetime.strptime(transaction_date, '%Y-%m-%d')

                    # Validate cost
                    cost_final_float = float(final_cost)
                    cost_pre_vat_float = cost_final_float / 1.24

                    # Validate status
                    valid_statuses = ["Εκκρεμεί", "Πληρώθηκε"]
                    if status not in valid_statuses:
                        raise ValueError(f"Κατάσταση '{status}' μη έγκυρη")

                    # Get or create customer
                    customer_id = db.get_customer_by_name(customer_name)
                    if not customer_id:
                        db.add_customer(customer_name)
                        customer_id = db.get_customer_by_name(customer_name)

                    # Add transaction
                    db.add_transaction(
                        customer_id, service_id, notes, transaction_date,
                        cost_pre_vat_float, cost_final_float, status
                    )

                    success_count += 1
                    log.append(f"✅ ΓΡΑΜΜΗ {row_idx}: Επιτυχία - {customer_name}")

                except Exception as e:
                    fail_count += 1
                    log.append(f"❌ ΓΡΑΜΜΗ {row_idx}: Σφάλμα - {str(e)}")

            # Summary
            summary = f"""
╔══════════════════════════════════════════╗
║      ΑΠΟΤΕΛΕΣΜΑΤΑ ΕΙΣΑΓΩΓΗΣ             ║
╠══════════════════════════════════════════╣
║  ✅ Επιτυχίες:  {success_count:4d}                     ║
║  ❌ Αποτυχίες:  {fail_count:4d}                     ║
╚══════════════════════════════════════════╝

ΛΕΠΤΟΜΕΡΕΙΕΣ:
{"=" * 50}
"""
            self.import_log_textbox.insert("1.0", summary + "\n".join(log))

            # Log the import
            db.add_audit_log(
                "IMPORT", "transactions", 0,
                f"Μαζική εισαγωγή: {success_count} επιτυχίες, {fail_count} αποτυχίες",
                "", ""
            )

            self.refresh_main_table()
            messagebox.showinfo("Ολοκλήρωση",
                              f"Η εισαγωγή ολοκληρώθηκε!\n\n✅ Επιτυχίες: {success_count}\n❌ Αποτυχίες: {fail_count}")

        except Exception as e:
            self.import_log_textbox.insert("1.0", f"❌ ΚΡΙΣΙΜΟ ΣΦΑΛΜΑ:\n{str(e)}")

        self.import_log_textbox.configure(state="disabled")

    # ========== LOG TAB ==========

    def create_log_tab(self):
        """Create the audit log viewer tab"""
        self.log_tab.grid_columnconfigure(0, weight=1)
        self.log_tab.grid_rowconfigure(2, weight=1)

        # Header
        header_frame = ctk.CTkFrame(self.log_tab)
        header_frame.grid(row=0, column=0, padx=20, pady=20, sticky="ew")

        title_label = ctk.CTkLabel(
            header_frame,
            text="📋 Ιστορικό Αλλαγών",
            font=ctk.CTkFont(size=22, weight="bold")
        )
        title_label.pack(side="left", pady=15, padx=20)

        refresh_btn = ctk.CTkButton(
            header_frame,
            text="🔄 Ανανέωση",
            command=self.refresh_audit_log,
            height=35,
            width=120
        )
        refresh_btn.pack(side="right", pady=15, padx=20)

        # Filters
        filter_frame = ctk.CTkFrame(self.log_tab)
        filter_frame.grid(row=1, column=0, padx=20, pady=(0, 10), sticky="ew")

        filter_label = ctk.CTkLabel(filter_frame, text="Φίλτρα:", font=ctk.CTkFont(weight="bold"))
        filter_label.pack(side="left", padx=(20, 10), pady=15)

        self.log_action_filter = ctk.StringVar(value="Όλα")
        action_menu = ctk.CTkOptionMenu(
            filter_frame,
            variable=self.log_action_filter,
            values=["Όλα", "INSERT", "UPDATE", "DELETE", "IMPORT"],
            command=lambda x: self.refresh_audit_log(),
            width=120
        )
        action_menu.pack(side="left", padx=5, pady=15)

        self.log_table_filter = ctk.StringVar(value="Όλα")
        table_menu = ctk.CTkOptionMenu(
            filter_frame,
            variable=self.log_table_filter,
            values=["Όλα", "transactions", "customers", "services"],
            command=lambda x: self.refresh_audit_log(),
            width=140
        )
        table_menu.pack(side="left", padx=5, pady=15)

        # Log Treeview
        tree_frame = ctk.CTkFrame(self.log_tab)
        tree_frame.grid(row=2, column=0, padx=20, pady=(0, 20), sticky="nsew")

        columns = ("ID", "Ενέργεια", "Πίνακας", "Περιγραφή", "Ημ/νία")
        self.log_tree = ttk.Treeview(tree_frame, columns=columns, show="headings")

        self.log_tree.heading("ID", text="ID")
        self.log_tree.heading("Ενέργεια", text="Ενέργεια")
        self.log_tree.heading("Πίνακας", text="Πίνακας")
        self.log_tree.heading("Περιγραφή", text="Περιγραφή")
        self.log_tree.heading("Ημ/νία", text="Ημερομηνία & Ώρα")

        self.log_tree.column("ID", width=60, anchor="center")
        self.log_tree.column("Ενέργεια", width=100, anchor="center")
        self.log_tree.column("Πίνακας", width=120)
        self.log_tree.column("Περιγραφή", width=500)
        self.log_tree.column("Ημ/νία", width=180, anchor="center")

        # Color coding
        self.log_tree.tag_configure('insert', background='#166534', foreground='white')
        self.log_tree.tag_configure('update', background='#1e40af', foreground='white')
        self.log_tree.tag_configure('delete', background='#991b1b', foreground='white')
        self.log_tree.tag_configure('import', background='#7c2d12', foreground='white')

        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.log_tree.yview)
        self.log_tree.configure(yscrollcommand=scrollbar.set)

        self.log_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        self.refresh_audit_log()

    def refresh_audit_log(self):
        """Refresh the audit log display"""
        # Clear existing items
        for item in self.log_tree.get_children():
            self.log_tree.delete(item)

        # Get filters
        action_filter = None if self.log_action_filter.get() == "Όλα" else self.log_action_filter.get()
        table_filter = None if self.log_table_filter.get() == "Όλα" else self.log_table_filter.get()

        # Get logs
        logs = db.get_audit_logs(limit=200, filter_action=action_filter, filter_table=table_filter)

        for log in logs:
            log_id, action, table, record_id, description, old_val, new_val, timestamp = log
            tag = action.lower()
            self.log_tree.insert("", "end", values=(log_id, action, table, description, timestamp), tags=(tag,))


# ========== RUN APPLICATION ==========

if __name__ == "__main__":
    app = App()
    app.mainloop()
