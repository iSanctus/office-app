# app.py (Final Version with Excel Batch Import)
import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog, TclError, Listbox
import database as db
import datetime
import shutil
import os
import csv
from openpyxl import Workbook, load_workbook # NEW: Import for Excel handling
from openpyxl.utils import get_column_letter

# ---- Pop-up window for editing transactions ----
class EditTransactionWindow(ctk.CTkToplevel):
    # (This class remains unchanged)
    def __init__(self, master, transaction_id):
        super().__init__(master)
        self.master_app = master; self.transaction_id = transaction_id; self.title("Επεξεργασία Συναλλαγής"); self.geometry("400x300"); self.transient(master); self.grab_set()
        _id, current_notes, current_status = db.get_transaction_details(self.transaction_id)
        ctk.CTkLabel(self, text="Κατάσταση Πληρωμής:", font=ctk.CTkFont(weight="bold")).pack(pady=(20, 5), padx=20, anchor="w")
        self.status_var = ctk.StringVar(value=current_status); self.status_menu = ctk.CTkOptionMenu(self, variable=self.status_var, values=["Εκκρεμεί", "Πληρώθηκε", "Καθυστέρηση"]); self.status_menu.pack(padx=20, fill="x")
        ctk.CTkLabel(self, text="Σχόλια / Παρατηρήσεις:", font=ctk.CTkFont(weight="bold")).pack(pady=(20, 5), padx=20, anchor="w")
        self.notes_textbox = ctk.CTkTextbox(self, height=100); self.notes_textbox.pack(padx=20, fill="both", expand=True); self.notes_textbox.insert("1.0", current_notes if current_notes else "")
        ctk.CTkButton(self, text="Αποθήκευση", command=self.save_changes, height=40).pack(pady=20, padx=20, fill="x")
    def save_changes(self):
        new_status = self.status_var.get(); new_notes = self.notes_textbox.get("1.0", "end-1c").strip()
        db.update_transaction(self.transaction_id, new_status, new_notes); messagebox.showinfo("Επιτυχία", "Οι αλλαγές αποθηκεύτηκαν.", parent=self)
        self.master_app.search_customer(); self.master_app.refresh_main_table(); self.destroy()

# --- Main Application Class ---
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Σύστημα Διαχείρισης Έργων v7.0")
        self.geometry("1300x750")
        db.connect_db()
        self.current_customer_records = []

        self.tab_view = ctk.CTkTabview(self); self.tab_view.pack(expand=True, fill="both", padx=10, pady=10)
        self.main_tab = self.tab_view.add("Κεντρική Σελίδα")
        self.customer_tab = self.tab_view.add("Αναζήτηση Πελάτη")
        self.services_tab = self.tab_view.add("Διαχείριση Υπηρεσιών")
        self.import_tab = self.tab_view.add("Μαζική Εισαγωγή") # NEW TAB

        self.create_main_tab()
        self.create_customer_search_tab()
        self.create_services_tab()
        self.create_import_tab() # NEW TAB
        
        self.tab_view.set("Κεντρική Σελίδα")
        self.bind("<Button-1>", self.on_global_click)
        
    # --- NEW: TAB FOR BATCH IMPORT ---
    def create_import_tab(self):
        self.import_tab.grid_columnconfigure(0, weight=1)
        self.import_tab.grid_rowconfigure(2, weight=1)

        info_frame = ctk.CTkFrame(self.import_tab)
        info_frame.grid(row=0, column=0, padx=20, pady=20, sticky="ew")

        ctk.CTkLabel(info_frame, text="Βήμα 1: Κατεβάστε το πρότυπο αρχείο Excel για να συμπληρώσετε τα δεδομένα σας.", wraplength=500).pack(pady=5)
        ctk.CTkButton(info_frame, text="Λήψη Προτύπου Excel (.xlsx)", command=self.download_template).pack(pady=10)
        
        ctk.CTkLabel(info_frame, text="Βήμα 2: Αφού συμπληρώσετε το αρχείο, εισάγετέ το στην εφαρμογή.", wraplength=500).pack(pady=5)
        ctk.CTkButton(info_frame, text="Εισαγωγή από Αρχείο Excel", command=self.import_from_excel).pack(pady=10)

        log_frame = ctk.CTkFrame(self.import_tab)
        log_frame.grid(row=1, column=0, padx=20, pady=10, sticky="ew")
        ctk.CTkLabel(log_frame, text="Αποτελέσματα Εισαγωγής:").pack(anchor="w")

        self.import_log_textbox = ctk.CTkTextbox(self.import_tab, wrap="word")
        self.import_log_textbox.grid(row=2, column=0, padx=20, pady=(0, 20), sticky="nsew")
        self.import_log_textbox.insert("end", "Εδώ θα εμφανιστούν τα αποτελέσματα της διαδικασίας εισαγωγής...")
        self.import_log_textbox.configure(state="disabled")

    def download_template(self):
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="Προτυπο_Εισαγωγης.xlsx",
            title="Αποθήκευση Προτύπου"
        )
        if not filepath: return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            
            headers = [
                'Ονοματεπώνυμο Πελάτη', 'Υπηρεσία', 'Ημερομηνία (YYYY-MM-DD)', 
                'Τελικό Κόστος (με ΦΠΑ)', 'Κατάσταση', 'Σχόλια'
            ]
            ws.append(headers)

            # Autofit column width for better readability
            for col_idx, header in enumerate(headers, 1):
                column_letter = get_column_letter(col_idx)
                ws.column_dimensions[column_letter].width = len(header) + 5
            
            wb.save(filepath)
            messagebox.showinfo("Επιτυχία", f"Το πρότυπο αποθηκεύτηκε με επιτυχία στο:\n{filepath}")
        except Exception as e:
            messagebox.showerror("Σφάλμα", f"Απέτυχε η δημιουργία του προτύπου.\nΣφάλμα: {e}")

    def import_from_excel(self):
        filepath = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Επιλογή Αρχείου Excel για Εισαγωγή"
        )
        if not filepath: return
        
        if not messagebox.askyesno("Επιβεβαίωση", "Είστε σίγουροι ότι θέλετε να ξεκινήσετε την εισαγωγή δεδομένων από αυτό το αρχείο;"):
            return

        self.import_log_textbox.configure(state="normal")
        self.import_log_textbox.delete("1.0", "end")
        
        log = []
        success_count = 0
        fail_count = 0
        
        try:
            wb = load_workbook(filepath, data_only=True)
            ws = wb.active
            
            # Create a dictionary of available services for quick lookup
            available_services = {name.lower(): sid for sid, name in db.get_services()}
            
            # Skip header row
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                try:
                    customer_name = str(row[0].value).strip() if row[0].value else None
                    service_name = str(row[1].value).strip() if row[1].value else None
                    date_val = row[2].value
                    final_cost = row[3].value
                    status = str(row[4].value).strip() if row[4].value else None
                    notes = str(row[5].value).strip() if row[5].value else ""

                    # --- Data Validation ---
                    if not all([customer_name, service_name, date_val, final_cost, status]):
                        raise ValueError("Λείπουν υποχρεωτικά δεδομένα.")
                    
                    # Validate service
                    service_id = available_services.get(service_name.lower())
                    if not service_id:
                        raise ValueError(f"Η υπηρεσία '{service_name}' δεν υπάρχει. Προσθέστε την πρώτα.")
                    
                    # Validate date
                    if isinstance(date_val, datetime.datetime):
                        transaction_date = date_val.strftime('%Y-%m-%d')
                    else:
                        transaction_date = str(date_val) # Assume it's a string in the correct format
                        datetime.datetime.strptime(transaction_date, '%Y-%m-%d') # This will raise error if format is wrong

                    # Validate cost
                    cost_final_float = float(final_cost)
                    cost_pre_vat_float = cost_final_float / 1.24
                    
                    # Validate status
                    valid_statuses = ["Εκκρεμεί", "Πληρώθηκε", "Καθυστέρηση"]
                    if status not in valid_statuses:
                        raise ValueError(f"Η κατάσταση '{status}' δεν είναι έγκυρη.")
                        
                    # --- Process Data ---
                    customer_id = db.get_customer_by_name(customer_name)
                    if not customer_id:
                        db.add_customer(customer_name)
                        customer_id = db.get_customer_by_name(customer_name)
                    
                    db.add_transaction(customer_id, service_id, notes, transaction_date, cost_pre_vat_float, cost_final_float, status)
                    success_count += 1
                    log.append(f"ΓΡΑΜΜΗ {row_idx}: ΕΠΙΤΥΧΙΑ - Προστέθηκε εγγραφή για τον πελάτη '{customer_name}'.")

                except Exception as e:
                    fail_count += 1
                    log.append(f"ΓΡΑΜΜΗ {row_idx}: ΣΦΑΛΜΑ - {e}")

            # Final Log Summary
            log.insert(0, f"--- ΑΠΟΤΕΛΕΣΜΑΤΑ ΕΙΣΑΓΩΓΗΣ ---\nΕπιτυχίες: {success_count}\nΑποτυχίες: {fail_count}\n---------------------------------\n")
            self.import_log_textbox.insert("1.0", "\n".join(log))
            self.refresh_main_table() # Refresh the main table to show new entries
            messagebox.showinfo("Ολοκλήρωση", f"Η διαδικασία εισαγωγής ολοκληρώθηκε.\nΕπιτυχίες: {success_count}\nΑποτυχίες: {fail_count}")

        except Exception as e:
            self.import_log_textbox.insert("1.0", f"ΚΡΙΣΙΜΟ ΣΦΑΛΜΑ: Δεν ήταν δυνατό το διάβασμα του αρχείου.\n{e}")
        
        self.import_log_textbox.configure(state="disabled")


    # --- ALL OTHER FUNCTIONS from the previous version remain here ---
    # They are unchanged, so I'm including them for completeness.
    def calculate_vat(self, *args):
        try:
            final_cost = float(self.final_cost_var.get())
            pre_vat_cost = final_cost / 1.24
            self.cost_pre_vat_entry.configure(state="normal"); self.cost_pre_vat_entry.delete(0, 'end'); self.cost_pre_vat_entry.insert(0, f"{pre_vat_cost:.2f}"); self.cost_pre_vat_entry.configure(state="readonly")
        except (ValueError, TclError):
            self.cost_pre_vat_entry.configure(state="normal"); self.cost_pre_vat_entry.delete(0, 'end'); self.cost_pre_vat_entry.configure(state="readonly")

    def create_main_tab(self):
        self.main_tab.grid_columnconfigure(1, weight=1); self.main_tab.grid_rowconfigure(0, weight=1)
        left_panel_frame = ctk.CTkFrame(self.main_tab, width=350); left_panel_frame.grid(row=0, column=0, padx=10, pady=10, sticky="ns"); left_panel_frame.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(left_panel_frame, text="Νέα Καταχώρηση", font=ctk.CTkFont(size=20, weight="bold")).grid(row=0, column=0, pady=(10, 20), padx=20)
        customer_frame = ctk.CTkFrame(left_panel_frame, fg_color="transparent"); customer_frame.grid(row=1, column=0, padx=20, pady=(0, 10), sticky="ew"); customer_frame.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(customer_frame, text="Όνομα Πελάτη*").grid(row=0, column=0, sticky="w")
        self.customer_name_entry = ctk.CTkEntry(customer_frame); self.customer_name_entry.grid(row=1, column=0, sticky="ew"); self.customer_name_entry.bind("<KeyRelease>", self.update_main_autocomplete)
        self.main_autocomplete_listbox = Listbox(customer_frame, bg="#2b2b2b", fg="white", highlightthickness=0, selectbackground="#347083"); self.main_autocomplete_listbox.grid(row=2, column=0, sticky="nsew"); self.main_autocomplete_listbox.bind("<<ListboxSelect>>", self.select_from_main_autocomplete); self.main_autocomplete_listbox.grid_forget()
        ctk.CTkLabel(left_panel_frame, text="Υπηρεσία*").grid(row=2, column=0, sticky="w", padx=20)
        self.service_var = ctk.StringVar(value="Επιλέξτε Υπηρεσία..."); self.service_menu = ctk.CTkOptionMenu(left_panel_frame, variable=self.service_var, values=[]); self.service_menu.grid(row=3, column=0, sticky="ew", padx=20, pady=(0, 10)); self.update_services_dropdown()
        ctk.CTkLabel(left_panel_frame, text="Σχόλια / Παρατηρήσεις").grid(row=4, column=0, sticky="w", padx=20)
        self.notes_entry = ctk.CTkEntry(left_panel_frame); self.notes_entry.grid(row=5, column=0, sticky="ew", padx=20, pady=(0, 10))
        ctk.CTkLabel(left_panel_frame, text="Τελικό Κόστος (με ΦΠΑ)*").grid(row=6, column=0, sticky="w", padx=20)
        self.final_cost_var = ctk.StringVar(); self.final_cost_var.trace_add("write", self.calculate_vat); self.cost_final_entry = ctk.CTkEntry(left_panel_frame, textvariable=self.final_cost_var); self.cost_final_entry.grid(row=7, column=0, sticky="ew", padx=20, pady=(0, 10))
        ctk.CTkLabel(left_panel_frame, text="Κόστος προ ΦΠΑ (αυτόματο)").grid(row=8, column=0, sticky="w", padx=20)
        self.cost_pre_vat_entry = ctk.CTkEntry(left_panel_frame, state="readonly"); self.cost_pre_vat_entry.grid(row=9, column=0, sticky="ew", padx=20, pady=(0, 10))
        ctk.CTkLabel(left_panel_frame, text="Κατάσταση Πληρωμής*").grid(row=10, column=0, sticky="w", padx=20)
        self.status_var = ctk.StringVar(value="Εκκρεμεί"); self.status_menu = ctk.CTkOptionMenu(left_panel_frame, variable=self.status_var, values=["Εκκρεμεί", "Πληρώθηκε", "Καθυστέρηση"]); self.status_menu.grid(row=11, column=0, sticky="ew", padx=20, pady=(0, 10))
        self.attachment_path = ctk.StringVar(); self.attachment_label = ctk.CTkLabel(left_panel_frame, text="Κανένα αρχείο επιλεγμένο", text_color="gray"); self.attachment_label.grid(row=12, column=0, pady=(10, 5), padx=20)
        ctk.CTkButton(left_panel_frame, text="Επισύναψη Αρχείου", command=self.select_file).grid(row=13, column=0, sticky="ew", padx=20)
        ctk.CTkButton(left_panel_frame, text="Προσθήκη Εγγραφής", command=self.add_transaction, height=40).grid(row=14, column=0, sticky="ew", padx=20, pady=20)
        table_frame = ctk.CTkFrame(self.main_tab); table_frame.grid(row=0, column=1, padx=(0,10), pady=10, sticky="nsew");
        table_frame.grid_rowconfigure(1, weight=1); table_frame.grid_columnconfigure(0, weight=1)
        filter_frame = ctk.CTkFrame(table_frame); filter_frame.grid(row=0, column=0, columnspan=2, padx=10, pady=10, sticky="ew")
        ctk.CTkLabel(filter_frame, text="Φίλτρο:").pack(side="left", padx=(10, 5))
        self.filter_var = ctk.StringVar(value="Όλα"); ctk.CTkOptionMenu(filter_frame, variable=self.filter_var, values=["Όλα", "Εκκρεμεί", "Πληρώθηκε", "Καθυστέρηση"], command=self.refresh_main_table).pack(side="left")
        self.open_attachment_button = ctk.CTkButton(filter_frame, text="Προβολή Επισυναπτόμενου", state="disabled", command=self.open_attachment); self.open_attachment_button.pack(side="right", padx=10)
        columns = ("ID", "Πελάτης", "Υπηρεσία", "Σχόλια", "Ημερομηνία", "Ποσό", "Κατάσταση"); self.tree = ttk.Treeview(table_frame, columns=columns, show="headings")
        for col in columns: self.tree.heading(col, text=col)
        self.tree.column("ID", width=50, anchor="center"); self.tree.column("Πελάτης", width=150); self.tree.column("Υπηρεσία", width=200); self.tree.column("Σχόλια", width=200); self.tree.column("Ποσό", anchor="e", width=100)
        self.tree.grid(row=1, column=0, sticky="nsew"); self.tree.bind("<<TreeviewSelect>>", self.on_row_select)
        self.tree.tag_configure('paid', background='#3a6b35', foreground='white'); self.tree.tag_configure('unpaid', background='#800f2f', foreground='white')
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview); self.tree.configure(yscrollcommand=scrollbar.set); scrollbar.grid(row=1, column=1, sticky="ns")
        self.refresh_main_table()
    def update_main_autocomplete(self, event):
        search_term = self.customer_name_entry.get();
        if len(search_term) < 1: self.main_autocomplete_listbox.grid_forget(); return
        results = db.search_customers_by_prefix(search_term); self.main_autocomplete_listbox.delete(0, 'end')
        if results:
            self.main_autocomplete_listbox.grid(row=2, column=0, sticky="nsew")
            for item in results: self.main_autocomplete_listbox.insert('end', item)
        else: self.main_autocomplete_listbox.grid_forget()
    def select_from_main_autocomplete(self, event):
        if not self.main_autocomplete_listbox.curselection(): return
        selected_name = self.main_autocomplete_listbox.get(self.main_autocomplete_listbox.curselection())
        self.customer_name_entry.delete(0, 'end'); self.customer_name_entry.insert(0, selected_name); self.main_autocomplete_listbox.grid_forget()
    def on_row_select(self, event):
        selected_items = self.tree.selection()
        if selected_items:
            transaction_id = self.tree.item(selected_items[0])['values'][0]; attachment_path = db.get_transaction_attachment(transaction_id)
            self.open_attachment_button.configure(state="normal" if attachment_path else "disabled")
        else: self.open_attachment_button.configure(state="disabled")
    def open_attachment(self):
        selected_items = self.tree.selection();
        if not selected_items: return
        transaction_id = self.tree.item(selected_items[0])['values'][0]; attachment_path = db.get_transaction_attachment(transaction_id)
        if attachment_path and os.path.exists(attachment_path):
            try: os.startfile(os.path.realpath(attachment_path))
            except Exception as e: messagebox.showerror("Σφάλμα", f"Δεν ήταν δυνατό το άνοιγμα του αρχείου:\n{e}")
        else: messagebox.showwarning("Προσοχή", "Το επισυναπτόμενο αρχείο δεν βρέθηκε.")
    def create_services_tab(self):
        self.services_tab.grid_columnconfigure(0, weight=1); self.services_tab.grid_rowconfigure(1, weight=1)
        add_service_frame = ctk.CTkFrame(self.services_tab); add_service_frame.grid(row=0, column=0, padx=20, pady=20, sticky="ew")
        ctk.CTkLabel(add_service_frame, text="Όνομα Νέας Υπηρεσίας:").pack(side="left", padx=10)
        self.new_service_entry = ctk.CTkEntry(add_service_frame, width=300); self.new_service_entry.pack(side="left", expand=True, fill="x", padx=10)
        ctk.CTkButton(add_service_frame, text="Προσθήκη Υπηρεσίας", command=self.add_new_service).pack(side="left", padx=10)
        service_list_frame = ctk.CTkFrame(self.services_tab); service_list_frame.grid(row=1, column=0, padx=20, pady=(0, 20), sticky="nsew"); service_list_frame.grid_columnconfigure(0, weight=1); service_list_frame.grid_rowconfigure(0, weight=1)
        self.service_list_tree = ttk.Treeview(service_list_frame, columns=("ID", "Όνομα Υπηρεσίας"), show="headings"); self.service_list_tree.heading("ID", text="ID"); self.service_list_tree.heading("Όνομα Υπηρεσίας", text="Όνομα Υπηρεσίας"); self.service_list_tree.column("ID", width=50, anchor="center"); self.service_list_tree.grid(row=0, column=0, sticky="nsew")
        ctk.CTkButton(service_list_frame, text="Διαγραφή Επιλεγμένης Υπηρεσίας", command=self.delete_selected_service).grid(row=1, column=0, pady=10)
        self.refresh_service_list()
    def add_new_service(self):
        service_name = self.new_service_entry.get().strip()
        if service_name: db.add_service(service_name); self.new_service_entry.delete(0, 'end'); self.refresh_service_list(); self.update_services_dropdown()
        else: messagebox.showwarning("Ελλιπή στοιχεία", "Το όνομα της υπηρεσίας δεν μπορεί να είναι κενό.")
    def delete_selected_service(self):
        selected_items = self.service_list_tree.selection()
        if not selected_items: messagebox.showwarning("Καμία Επιλογή", "Παρακαλώ επιλέξτε μια υπηρεσία για διαγραφή."); return
        if messagebox.askyesno("Επιβεβαίωση", "Είστε σίγουροι ότι θέλετε να διαγράψετε αυτή την υπηρεσία; Η ενέργεια αυτή δεν αναιρείται και οι συναλλαγές θα δείχνουν 'Διαγραμμένη Υπηρεσία'."):
            for item in selected_items: db.delete_service(self.service_list_tree.item(item)['values'][0])
            self.refresh_service_list(); self.update_services_dropdown()
    def refresh_service_list(self):
        for item in self.service_list_tree.get_children(): self.service_list_tree.delete(item)
        for service in db.get_services(): self.service_list_tree.insert("", "end", values=service)
    def update_services_dropdown(self):
        services = db.get_services(); service_names = [s[1] for s in services] or ["-"]
        self.service_menu.configure(values=service_names); self.service_var.set(service_names[0] if service_names[0] != "-" else "Προσθέστε υπηρεσίες")
    def select_file(self):
        filepath = filedialog.askopenfilename()
        if filepath: self.attachment_path.set(filepath); self.attachment_label.configure(text=os.path.basename(filepath), text_color="white")
    def add_transaction(self):
        customer_name = self.customer_name_entry.get().strip(); service_name = self.service_var.get(); notes = self.notes_entry.get().strip(); cost_final = self.cost_final_entry.get(); cost_pre_vat = self.cost_pre_vat_entry.get(); status = self.status_var.get()
        if not all([customer_name, cost_final, service_name not in ["Επιλέξτε Υπηρεσία...", "Προσθέστε υπηρεσίες", "-"]]): messagebox.showerror("Σφάλμα", "Παρακαλώ συμπληρώστε όλα τα υποχρεωτικά πεδία (*)."); return
        customer_id = db.get_customer_by_name(customer_name)
        if not customer_id: db.add_customer(customer_name); customer_id = db.get_customer_by_name(customer_name)
        service_id = {name: sid for sid, name in db.get_services()}.get(service_name)
        final_attachment_path = ""; original_path = self.attachment_path.get()
        if original_path: filename = f"{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}_{os.path.basename(original_path)}"; final_attachment_path = os.path.join(db.ATTACHMENTS_DIR, filename); shutil.copy(original_path, final_attachment_path)
        db.add_transaction(customer_id, service_id, notes, datetime.date.today().strftime('%Y-%m-%d'), float(cost_pre_vat), float(cost_final), status, final_attachment_path)
        messagebox.showinfo("Επιτυχία", "Η εγγραφή προστέθηκε."); self.clear_form(); self.refresh_main_table()
    def clear_form(self):
        self.customer_name_entry.delete(0, 'end'); self.notes_entry.delete(0, 'end'); self.cost_final_entry.delete(0, 'end'); self.attachment_path.set(""); self.attachment_label.configure(text="Κανένα αρχείο επιλεγμένο", text_color="gray")
    def refresh_main_table(self, filter_choice=None):
        if filter_choice is None: filter_choice = self.filter_var.get()
        for item in self.tree.get_children(): self.tree.delete(item)
        records = db.get_all_transactions(filter_choice)
        for record in records:
            display_values = list(record); amount = record[5]; display_values[5] = f"{amount:.2f} €"; status = record[6]; tag = 'paid' if status == 'Πληρώθηκε' else 'unpaid'
            self.tree.insert("", "end", values=display_values, tags=(tag,))
    def create_customer_search_tab(self):
        self.customer_tab.grid_columnconfigure(0, weight=1); self.customer_tab.grid_rowconfigure(2, weight=1)
        search_frame = ctk.CTkFrame(self.customer_tab); search_frame.grid(row=0, column=0, padx=20, pady=10, sticky="ew"); search_frame.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(search_frame, text="Όνομα Πελάτη:").grid(row=0, column=0, padx=10, sticky="w")
        self.search_customer_entry = ctk.CTkEntry(search_frame, width=300); self.search_customer_entry.grid(row=1, column=0, padx=10, sticky="ew"); self.search_customer_entry.bind("<KeyRelease>", self.update_search_autocomplete)
        self.search_autocomplete_listbox = Listbox(search_frame, bg="#2b2b2b", fg="white", highlightthickness=0, selectbackground="#347083"); self.search_autocomplete_listbox.grid(row=2, column=0, padx=10, sticky="nsew"); self.search_autocomplete_listbox.bind("<<ListboxSelect>>", self.select_from_search_autocomplete); self.search_autocomplete_listbox.grid_forget()
        ctk.CTkButton(search_frame, text="Αναζήτηση", command=self.search_customer).grid(row=1, column=1, padx=10)
        results_frame = ctk.CTkFrame(self.customer_tab); results_frame.grid(row=2, column=0, padx=20, pady=(0,20), sticky="nsew")
        results_frame.grid_columnconfigure(0, weight=1); results_frame.grid_rowconfigure(2, weight=1)
        info_frame = ctk.CTkFrame(results_frame); info_frame.grid(row=0, column=0, sticky="ew", pady=(0, 5)); info_frame.grid_columnconfigure(0, weight=1)
        self.customer_info_label = ctk.CTkLabel(info_frame, text="Αναζητήστε έναν πελάτη...", font=ctk.CTkFont(size=16)); self.customer_info_label.grid(row=0, column=0, padx=10, sticky="w")
        self.paid_sum_label = ctk.CTkLabel(info_frame, text="", font=ctk.CTkFont(size=14, weight="bold"), text_color="#5cb85c"); self.paid_sum_label.grid(row=0, column=1, padx=10, sticky="e")
        self.unpaid_sum_label = ctk.CTkLabel(info_frame, text="", font=ctk.CTkFont(size=14, weight="bold"), text_color="#d9534f"); self.unpaid_sum_label.grid(row=0, column=2, padx=10, sticky="e")
        self.export_csv_button = ctk.CTkButton(info_frame, text="Εξαγωγή σε CSV", state="disabled", command=self.export_to_csv); self.export_csv_button.grid(row=0, column=3, padx=10)
        ctk.CTkLabel(results_frame, text="Διπλό κλικ σε μια εγγραφή για επεξεργασία", text_color="gray").grid(row=1, column=0, padx=10, pady=(0,5), sticky="w")
        customer_columns = ("ID", "Υπηρεσία", "Σχόλια", "Ημερομηνία", "Ποσό", "Κατάσταση"); self.customer_results_tree = ttk.Treeview(results_frame, columns=customer_columns, displaycolumns=["Υπηρεσία", "Σχόλια", "Ημερομηνία", "Ποσό", "Κατάσταση"], show="headings")
        for col in customer_columns[1:]: self.customer_results_tree.heading(col, text=col)
        self.customer_results_tree.grid(row=2, column=0, sticky="nsew"); self.customer_results_tree.bind("<Double-1>", self.on_double_click_customer_tree)
        self.customer_results_tree.tag_configure('paid', background='#3a6b35', foreground='white'); self.customer_results_tree.tag_configure('unpaid', background='#800f2f', foreground='white')
        scrollbar_cust = ttk.Scrollbar(results_frame, orient="vertical", command=self.customer_results_tree.yview); self.customer_results_tree.configure(yscrollcommand=scrollbar_cust.set); scrollbar_cust.grid(row=2, column=1, sticky="ns")
    def export_to_csv(self):
        if not self.current_customer_records: messagebox.showwarning("Δεν υπάρχουν δεδομένα", "Δεν υπάρχουν δεδομένα για εξαγωγή."); return
        customer_name = self.search_customer_entry.get().strip().replace(" ", "_"); suggested_filename = f"Αναφορά_{customer_name}_{datetime.date.today()}.csv"
        filepath = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv"), ("All files", "*.*")], initialfile=suggested_filename, title="Αποθήκευση αναφοράς ως CSV")
        if not filepath: return
        try:
            with open(filepath, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile); writer.writerow(['Υπηρεσία', 'Σχόλια', 'Ημερομηνία', 'Τελικό Ποσό', 'Κατάσταση'])
                for record in self.current_customer_records: writer.writerow([record[1], record[2], record[3], record[4], record[5]])
            messagebox.showinfo("Επιτυχία", f"Η αναφορά αποθηκεύτηκε με επιτυχία στο:\n{filepath}")
        except Exception as e: messagebox.showerror("Σφάλμα", f"Απέτυχε η αποθήκευση του αρχείου.\nΣφάλμα: {e}")
    def update_search_autocomplete(self, event):
        search_term = self.search_customer_entry.get();
        if len(search_term) < 1: self.search_autocomplete_listbox.grid_forget(); return
        results = db.search_customers_by_prefix(search_term); self.search_autocomplete_listbox.delete(0, 'end')
        if results:
            self.search_autocomplete_listbox.grid(row=2, column=0, padx=10, sticky="nsew")
            for item in results: self.search_autocomplete_listbox.insert('end', item)
        else: self.search_autocomplete_listbox.grid_forget()
    def select_from_search_autocomplete(self, event):
        if not self.search_autocomplete_listbox.curselection(): return
        selected_name = self.search_autocomplete_listbox.get(self.search_autocomplete_listbox.curselection())
        self.search_customer_entry.delete(0, 'end'); self.search_customer_entry.insert(0, selected_name); self.hide_autocomplete(); self.search_customer()
    def on_double_click_customer_tree(self, event):
        selected_item = self.customer_results_tree.focus();
        if not selected_item: return
        item_values = self.customer_results_tree.item(selected_item, 'values'); transaction_id = item_values[0]; EditTransactionWindow(self, transaction_id)
    def search_customer(self):
        customer_name = self.search_customer_entry.get().strip(); self.hide_autocomplete();
        if not customer_name: return
        for item in self.customer_results_tree.get_children(): self.customer_results_tree.delete(item)
        self.paid_sum_label.configure(text=""); self.unpaid_sum_label.configure(text=""); self.export_csv_button.configure(state="disabled"); self.current_customer_records = []
        records = db.get_transactions_by_customer(customer_name); self.current_customer_records = records
        if not records:
            self.customer_info_label.configure(text=f"Δεν βρέθηκαν υπηρεσίες για: '{customer_name}'"); messagebox.showinfo("Κανένα Αποτέλεσμα", f"Δεν βρέθηκαν εγγραφές για τον πελάτη '{customer_name}'.")
        else:
            self.customer_info_label.configure(text=f"Εμφάνιση υπηρεσιών για: {customer_name}", font=ctk.CTkFont(size=16, weight="bold")); self.export_csv_button.configure(state="normal")
            total_paid = 0; total_unpaid = 0
            for record in records:
                cost = record[4]; status = record[5]
                if status == 'Πληρώθηκε': total_paid += cost
                else: total_unpaid += cost
                display_values = list(record); display_values[4] = f"{cost:.2f} €"; tag = 'paid' if status == 'Πληρώθηκε' else 'unpaid'
                self.customer_results_tree.insert("", "end", values=display_values, tags=(tag,))
            self.paid_sum_label.configure(text=f"Πληρωμένα: {total_paid:.2f} €"); self.unpaid_sum_label.configure(text=f"Οφειλές: {total_unpaid:.2f} €")
    def on_global_click(self, event):
        widget = self.winfo_containing(event.x_root, event.y_root)
        if widget != self.main_autocomplete_listbox and widget != self.customer_name_entry: self.main_autocomplete_listbox.grid_forget()
        if widget != self.search_autocomplete_listbox and widget != self.search_customer_entry: self.search_autocomplete_listbox.grid_forget()
    def hide_autocomplete(self):
        self.main_autocomplete_listbox.grid_forget(); self.search_autocomplete_listbox.grid_forget()
        
if __name__ == "__main__":
    app = App()
    app.mainloop()